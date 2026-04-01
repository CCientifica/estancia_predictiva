import streamlit as st
import pandas as pd
import joblib
import os
import io
import plotly.express as px
from streamlit_option_menu import option_menu
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font

# =========================================================
# 1. CONFIGURACIÓN DE PÁGINA Y DISEÑO
# =========================================================
try:
    if os.path.exists("Logo_Clinica.png"):
        st.set_page_config(
            page_title="Predictor LOS | Clínica Sagrado Corazón",
            layout="wide",
            page_icon="Logo_Clinica.png"
        )
    elif os.path.exists("logo.jpg"):
        st.set_page_config(
            page_title="Predictor LOS | Clínica Sagrado Corazón",
            layout="wide",
            page_icon="logo.jpg"
        )
    else:
        st.set_page_config(
            page_title="Predictor LOS | Clínica Sagrado Corazón",
            layout="wide"
        )
except Exception:
    st.set_page_config(
        page_title="Predictor LOS | Clínica Sagrado Corazón",
        layout="wide"
    )

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Open+Sans:wght@400;600&family=Poppins:wght@500;700&family=Roboto:wght@400;500&display=swap');

    .stApp {
        background-color: #f4f6f9;
        font-family: 'Open Sans', sans-serif;
    }

    h1, h2, h3 {
        color: #253d5b;
        font-family: 'Poppins', sans-serif;
        font-weight: 700;
    }

    p, span, div, label {
        color: #333333;
        font-family: 'Roboto', sans-serif;
    }

    .stButton>button {
        background-color: #4e6c9f;
        color: white;
        border-radius: 6px;
        border: none;
        font-weight: 600;
        font-family: 'Open Sans', sans-serif;
        padding: 10px 24px;
        transition: all 0.3s ease;
    }

    .stButton>button:hover {
        background-color: #253d5b;
        color: white;
        box-shadow: 0 4px 8px rgba(0,0,0,0.15);
    }

    div[data-testid="metric-container"] {
        background-color: white;
        border: 1px solid #b6b5af;
        padding: 15px;
        border-radius: 8px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
    }
</style>
""", unsafe_allow_html=True)

# =========================================================
# 2. CARGA DEL MODELO
# =========================================================
@st.cache_resource
def cargar_modelo():
    if os.path.exists('modelo_estancia_v4.pkl') and os.path.exists('columnas_modelo_v4.pkl'):
        return joblib.load('modelo_estancia_v4.pkl'), joblib.load('columnas_modelo_v4.pkl'), "V4"
    elif os.path.exists('modelo_estancia_v3.pkl') and os.path.exists('columnas_modelo_v3.pkl'):
        return joblib.load('modelo_estancia_v3.pkl'), joblib.load('columnas_modelo_v3.pkl'), "V3"
    elif os.path.exists('modelo_estancia_v2.pkl') and os.path.exists('columnas_modelo_v2.pkl'):
        return joblib.load('modelo_estancia_v2.pkl'), joblib.load('columnas_modelo_v2.pkl'), "V2"
    else:
        return joblib.load('modelo_estancia_v1.pkl'), joblib.load('columnas_modelo.pkl'), "V1"

try:
    modelo, columnas_entrenamiento, version_modelo = cargar_modelo()
except Exception as e:
    st.error(f"Error cargando el modelo: {e}")
    st.stop()

# =========================================================
# 3. FUNCIONES AUXILIARES
# =========================================================
def normalizar_texto(valor):
    if pd.isna(valor):
        return ""
    return str(valor).strip().upper()

def leer_archivo_subido(archivo_subido):
    nombre = archivo_subido.name.lower()

    if nombre.endswith(".csv"):
        try:
            return pd.read_csv(archivo_subido, sep=';', encoding='latin1')
        except Exception:
            archivo_subido.seek(0)
            return pd.read_csv(archivo_subido)
    return pd.read_excel(archivo_subido)

def estandarizar_columnas(df):
    columnas_nuevas = []

    for col in df.columns:
        c = str(col).strip().upper()

        if c == 'EDAD':
            columnas_nuevas.append('edad')
        elif c in ['FECINGRESO', 'FECHA INGRESO', 'FECHA_DE_INGRESO', 'FECHAINGRESO']:
            columnas_nuevas.append('FechaIngreso')
        elif c in ['PABACTUAL', 'PABELLON ACTUAL', 'PABELLONACTUAL', 'SERVICIO ACTUAL', 'UBICACION ACTUAL']:
            columnas_nuevas.append('PabellonActual')
        elif c in ['PABINGRESO', 'PABELLON DE INGRESO', 'PABELLONINGRESO', 'SERVICIO INGRESO', 'UBICACION INGRESO']:
            columnas_nuevas.append('PabellonIngresoOriginal')
        elif c in ['ESPECTRATANTE', 'ESPECIALIDAD TRATANTE', 'ESP', 'ESPECIALIDAD']:
            columnas_nuevas.append('Esp')
        elif c in ['DX1', 'COD. DIAG.', 'COD. DIAG', 'DX', 'CODIGO DIAGNOSTICO', 'COD DIAG']:
            columnas_nuevas.append('Dx')
        elif c in ['DIAGNOSTICO', 'DX1NOMBRE', 'DX1 NOMBRE', 'DESCRIPCION DIAGNOSTICO', 'NOMBRE DIAGNOSTICO']:
            columnas_nuevas.append('DxNombrePrincipal')
        elif c in ['DX2NOMBRE', 'DX2 NOMBRE', 'DIAGNOSTICO 2']:
            columnas_nuevas.append('Dx2Nombre')
        elif c in ['DX3NOMBRE', 'DX3 NOMBRE', 'DIAGNOSTICO 3']:
            columnas_nuevas.append('Dx3Nombre')
        elif c == 'CAMA':
            columnas_nuevas.append('CAMA')
        elif c == 'SEXO':
            columnas_nuevas.append('Sexo')
        elif c in ['FECHANACIMIENTO', 'FECHA DE NACIMIENTO', 'FECNACIMIENTO']:
            columnas_nuevas.append('FechaNacimiento')
        elif c in ['IDENTIFICACION', 'NUM DOC', 'TFCEDU', 'NOINGRESO', 'NUMERO DOCUMENTO', 'NRO DOCUMENTO']:
            columnas_nuevas.append('Identificacion')
        elif c in ['PACIENTE', 'NOMBRE1', 'NOMBRES', 'NOMBRE']:
            columnas_nuevas.append('Paciente_Nombre')
        elif c in ['APELLIDO1', 'APELLIDOS', 'APELLIDO']:
            columnas_nuevas.append('Paciente_Apellido')
        else:
            columnas_nuevas.append(str(col).strip())

    df.columns = columnas_nuevas
    return df

def clasificar_ubicacion_visual_desde_cama(cama):
    if pd.isna(cama):
        return "DESCONOCIDO"

    cama = str(cama).strip().upper()

    if cama in [str(i) for i in range(201, 223)]:
        return "SEGUNDO PISO (HEMATO-ONCOLOGIA)"

    tercer_piso = [str(i) for i in range(306, 323)] + [
        "301PA", "301PB", "302P", "303P", "304P", "305P",
        "323P", "324P", "325P", "326P"
    ]
    if cama in tercer_piso:
        return "TERCER PISO"

    cuarto_piso = [str(i) for i in range(401, 426)] + [
        "416A", "416B", "417A", "417B", "418A", "418B", "419A", "419B",
        "420A", "420B", "423A", "423B", "424A", "424B", "425A", "425B"
    ]
    if cama in cuarto_piso:
        return "CUARTO PISO"

    if cama.startswith("UCI"):
        return "UCI"
    if cama.startswith("UCE"):
        return "UCE"
    if cama.startswith(("AU", "PAS", "PED", "REA", "SI", "YES", "H2")):
        return "URGENCIAS"

    return "DESCONOCIDO"

def homologar_pabellon_modelo(texto):
    """
    Convierte ubicaciones reales / textos libres a categorías más cercanas
    a las que razonablemente usó el modelo en entrenamiento.
    """
    txt = normalizar_texto(texto)

    if txt == "":
        return "HOSPITALIZACION"

    # Críticos
    if any(k in txt for k in ["UCI", "CUIDADO CRITICO", "CUIDADO CRÍTICO", "CRITICO", "CRÍTICO", "UCE"]):
        return "UCI"

    # Urgencias / observación / triage
    if any(k in txt for k in ["URGEN", "OBSERV", "TRIAGE", "REA", "PASILLO", "YES", "AU", "PED"]):
        return "URGENCIAS"

    # Quirúrgicos
    if any(k in txt for k in ["CIRUG", "QUIR", "SALA DE CIRUGIA", "SALA DE CIRUGÍA"]):
        return "CIRUGIA"

    # Pisos / hospitalización
    if any(k in txt for k in [
        "PISO", "HOSPITAL", "MEDICINA", "HEMATO", "ONCO", "ORTO", "GINECO",
        "CUARTO PISO", "TERCER PISO", "SEGUNDO PISO"
    ]):
        return "HOSPITALIZACION"

    return "HOSPITALIZACION"

def preparar_pabellones(df):
    # Ubicación visual real
    if 'PabellonActual' not in df.columns:
        df['PabellonActual'] = ""

    if 'PabellonIngresoOriginal' not in df.columns:
        df['PabellonIngresoOriginal'] = ""

    if 'CAMA' in df.columns:
        ubicacion_por_cama = df['CAMA'].apply(clasificar_ubicacion_visual_desde_cama)
    else:
        ubicacion_por_cama = pd.Series(["DESCONOCIDO"] * len(df), index=df.index)

    # Prioridad para visualización: actual > ingreso > cama
    df['PabellonVisual'] = (
        df['PabellonActual'].fillna("").astype(str).str.strip()
    )
    mascara_vacia_visual = df['PabellonVisual'].eq("")
    df.loc[mascara_vacia_visual, 'PabellonVisual'] = (
        df.loc[mascara_vacia_visual, 'PabellonIngresoOriginal'].fillna("").astype(str).str.strip()
    )

    mascara_vacia_visual = df['PabellonVisual'].eq("")
    df.loc[mascara_vacia_visual, 'PabellonVisual'] = ubicacion_por_cama[mascara_vacia_visual]

    df['PabellonVisual'] = df['PabellonVisual'].fillna("DESCONOCIDO").astype(str).str.upper()

    # Campo específico para el modelo
    fuente_modelo = df['PabellonActual'].fillna("").astype(str).str.strip()
    mascara_vacia_modelo = fuente_modelo.eq("")
    fuente_modelo.loc[mascara_vacia_modelo] = df['PabellonIngresoOriginal'].fillna("").astype(str).str.strip()[mascara_vacia_modelo]

    mascara_vacia_modelo = fuente_modelo.eq("")
    fuente_modelo.loc[mascara_vacia_modelo] = ubicacion_por_cama[mascara_vacia_modelo]

    df['PabellonIngreso'] = fuente_modelo.apply(homologar_pabellon_modelo)

    return df

def preparar_edad(df):
    if 'edad' in df.columns:
        df['edad'] = df['edad'].astype(str).str.extract(r'(\d+(?:\.\d+)?)')[0]
        df['edad'] = pd.to_numeric(df['edad'], errors='coerce').fillna(50)
    elif 'FechaNacimiento' in df.columns:
        df['FechaNacimiento'] = pd.to_datetime(df['FechaNacimiento'], errors='coerce')
        df['edad'] = (pd.Timestamp.now() - df['FechaNacimiento']).dt.total_seconds() / (3600 * 24 * 365.25)
        df['edad'] = df['edad'].fillna(50)
    else:
        df['edad'] = 50
        st.toast(
            "⚠️ El censo no contenía la edad de los pacientes. El sistema asumió 50 años como promedio preventivo.",
            icon="⚠️"
        )
    return df

def preparar_campos_basicos(df):
    if 'Sexo' not in df.columns:
        df['Sexo'] = 'M'
    df['Sexo'] = df['Sexo'].fillna('M').astype(str).str.strip().str.upper()
    df.loc[~df['Sexo'].isin(['M', 'F']), 'Sexo'] = 'M'

    if 'Esp' not in df.columns:
        df['Esp'] = 'MEDICINA INTERNA'
    df['Esp'] = df['Esp'].fillna('MEDICINA INTERNA').astype(str).str.strip().str.upper()

    if 'Dx' not in df.columns:
        df['Dx'] = ""
    df['Dx'] = df['Dx'].fillna("").astype(str).str.strip().str.upper()

    if 'DxNombrePrincipal' not in df.columns:
        df['DxNombrePrincipal'] = ""
    if 'Dx2Nombre' not in df.columns:
        df['Dx2Nombre'] = ""
    if 'Dx3Nombre' not in df.columns:
        df['Dx3Nombre'] = ""

    df['DxNombrePrincipal'] = df['DxNombrePrincipal'].fillna("").astype(str).str.upper()
    df['Dx2Nombre'] = df['Dx2Nombre'].fillna("").astype(str).str.upper()
    df['Dx3Nombre'] = df['Dx3Nombre'].fillna("").astype(str).str.upper()

    return df

def preparar_dias_actuales(df):
    if 'FechaIngreso' in df.columns:
        df['FechaIngreso'] = pd.to_datetime(df['FechaIngreso'], errors='coerce')
        df['Dias_Actuales'] = (pd.Timestamp.now() - df['FechaIngreso']).dt.total_seconds() / (24 * 3600)
        df['Dias_Actuales'] = df['Dias_Actuales'].round(1).fillna(0)
    else:
        df['FechaIngreso'] = pd.NaT
        df['Dias_Actuales'] = 0
    return df

def construir_texto_dx(df):
    piezas = [
        df.get('Dx', pd.Series("", index=df.index)).fillna("").astype(str).str.upper(),
        df.get('DxNombrePrincipal', pd.Series("", index=df.index)).fillna("").astype(str).str.upper(),
        df.get('Dx2Nombre', pd.Series("", index=df.index)).fillna("").astype(str).str.upper(),
        df.get('Dx3Nombre', pd.Series("", index=df.index)).fillna("").astype(str).str.upper(),
    ]
    texto = piezas[0]
    for serie in piezas[1:]:
        texto = texto + " " + serie
    return texto.str.replace(r"\s+", " ", regex=True).str.strip()

def construir_variables_modelo(df, version_modelo_local):
    if version_modelo_local in ["V2", "V3", "V4"]:
        df['Texto_Dx'] = construir_texto_dx(df)

        df['Diabetes'] = df['Texto_Dx'].str.contains(r'DIABETES', na=False).astype(int)
        df['Hipertension'] = df['Texto_Dx'].str.contains(r'HIPERTENSION|HIPERTENSI[ÓO]N|HTA|PRESION ALTA', na=False).astype(int)
        df['Cardiaca'] = df['Texto_Dx'].str.contains(r'CARDIAC|INFARTO|ISQUEMI|FALLA CARDIACA|INSUFICIENCIA CARDIACA', na=False).astype(int)
        df['EPOC'] = df['Texto_Dx'].str.contains(r'EPOC|PULMONAR OBSTRUCTIVA', na=False).astype(int)
        df['Hemato_Onco'] = df['Texto_Dx'].str.contains(r'CANCER|CÁNCER|TUMOR|LEUCEMIA|LINFOMA|NEOPLASIA|MALIGN|HEMATO|ONCO', na=False).astype(int)
        df['Quimio'] = df['Texto_Dx'].str.contains(r'QUIMIO|QUIMIOTERAP', na=False).astype(int)
        df['Hemofilia'] = df['Texto_Dx'].str.contains(r'HEMOFILIA', na=False).astype(int)
        df['Porfiria'] = df['Texto_Dx'].str.contains(r'PORFIRIA', na=False).astype(int)
        df['Renal'] = df['Texto_Dx'].str.contains(r'RENAL|NEFRO|IRC|INSUFICIENCIA RENAL', na=False).astype(int)
        df['VIH'] = df['Texto_Dx'].str.contains(r'VIH|SIDA|INMUNODEFICIENCIA HUMANA', na=False).astype(int)

        cols_enfermedades = [
            'Diabetes', 'Hipertension', 'Cardiaca', 'EPOC', 'Hemato_Onco',
            'Quimio', 'Hemofilia', 'Porfiria', 'Renal', 'VIH'
        ]
        df['Total_Comorbilidades'] = df[cols_enfermedades].sum(axis=1)

        digits = 4 if version_modelo_local in ["V3", "V4"] else 3
        df['Dx_Agrupado'] = df['Dx'].astype(str).str[:digits]
        df['Dx_Agrupado'] = df['Dx_Agrupado'].replace("", "DESCONOCIDO")

        columnas_x = [
            'edad', 'Sexo', 'PabellonIngreso', 'Esp', 'Dx_Agrupado',
            'Diabetes', 'Hipertension', 'Cardiaca', 'EPOC', 'Hemato_Onco',
            'Quimio', 'Hemofilia', 'Porfiria', 'Renal', 'VIH', 'Total_Comorbilidades'
        ]
    else:
        df['Dx_Agrupado'] = df['Dx'].astype(str).str[:3].replace("", "DESCONOCIDO")
        if 'Complejidad' not in df.columns:
            df['Complejidad'] = 0
        columnas_x = ['edad', 'Sexo', 'PabellonIngreso', 'Esp', 'Dx_Agrupado', 'Complejidad']

    return df, columnas_x

def limitar_predicciones_v4(df):
    """
    No capamos ni manipulamos fuerte el modelo.
    Solo evitamos cosas absurdas como negativos o casi cero.
    """
    df['Prediccion_Estancia'] = pd.to_numeric(df['Prediccion_Estancia'], errors='coerce')
    df.loc[df['Prediccion_Estancia'] < 0.5, 'Prediccion_Estancia'] = 0.5
    return df

def auditar_predicciones(df, version_modelo_local):
    alertas = []
    riesgo_prolongado = 0

    if version_modelo_local == "V4":
        df = limitar_predicciones_v4(df)

        df['Prediccion_Estancia_Texto'] = df['Prediccion_Estancia'].apply(
            lambda x: f"{x:.1f} días" if pd.notna(x) else "No disponible"
        )

        def categorizar_estancia(x):
            if pd.isna(x):
                return "Sin predicción"
            if x < 3:
                return "Estancia Corta (<3 días)"
            elif x <= 7:
                return "Estancia Media (3 a 7 días)"
            return "Estancia Prolongada (>7 días)"

        df['Categoria_Estancia'] = df['Prediccion_Estancia'].apply(categorizar_estancia)

        for _, row in df.iterrows():
            dias_pred = row.get('Prediccion_Estancia', None)
            dias_actuales = row.get('Dias_Actuales', 0)

            if pd.notna(dias_pred) and dias_pred > 7:
                riesgo_prolongado += 1

            if pd.notna(dias_pred) and dias_actuales > dias_pred:
                alertas.append("Desviado - Límite Superado")
            else:
                alertas.append("Normal - Dentro del límite estimado")
    else:
        df['Prediccion_Estancia_Texto'] = df['Prediccion_Estancia'].astype(str)
        df['Categoria_Estancia'] = df['Prediccion_Estancia'].astype(str)

        for _, row in df.iterrows():
            pred = str(row.get('Prediccion_Estancia', ''))
            dias_actuales = row.get('Dias_Actuales', 0)

            if "PROLONGADA" in pred.upper():
                riesgo_prolongado += 1

            if "CORTA" in pred.upper() and dias_actuales > 3:
                alertas.append("Desviado - Límite Superado")
            elif "MEDIA" in pred.upper() and dias_actuales > 7:
                alertas.append("Desviado - Límite Superado")
            else:
                alertas.append("Normal - Dentro del límite estimado")

    df['Estado_Auditoria'] = alertas
    return df, riesgo_prolongado, alertas

def exportar_excel_estilizado(df_mostrar):
    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_mostrar.to_excel(writer, index=False, sheet_name='Censo_Auditable')
        worksheet = writer.sheets['Censo_Auditable']

        # Encabezados
        fill_header = PatternFill(fill_type="solid", fgColor="253d5b")
        font_header = Font(color="FFFFFF", bold=True)

        for cell in worksheet[1]:
            cell.fill = fill_header
            cell.font = font_header

        # Ajuste de ancho
        for idx, col in enumerate(df_mostrar.columns, start=1):
            max_len = max(
                df_mostrar[col].astype(str).map(len).max() if len(df_mostrar) > 0 else 0,
                len(str(col))
            ) + 2
            worksheet.column_dimensions[get_column_letter(idx)].width = min(max_len, 40)

        # Resaltado de desviados
        estado_col_idx = None
        for idx, col in enumerate(df_mostrar.columns, start=1):
            if col == 'Estado_Auditoria':
                estado_col_idx = idx
                break

        if estado_col_idx is not None:
            fill_rojo = PatternFill(fill_type="solid", fgColor="FFDCE0")
            for row_idx in range(2, len(df_mostrar) + 2):
                estado_valor = worksheet.cell(row=row_idx, column=estado_col_idx).value
                if estado_valor and "Desviado" in str(estado_valor):
                    for col_idx in range(1, len(df_mostrar.columns) + 1):
                        worksheet.cell(row=row_idx, column=col_idx).fill = fill_rojo

    return buffer.getvalue()

# =========================================================
# 4. ENCABEZADO
# =========================================================
col_logo, col_tit = st.columns([1, 6])

with col_logo:
    if os.path.exists("Logo_Clinica.png"):
        st.image("Logo_Clinica.png", width="stretch")
    elif os.path.exists("logo.jpg"):
        st.image("logo.jpg", width="stretch")

with col_tit:
    st.title("Sistema de Alerta Temprana y Gestión de Camas")
    st.markdown("Plataforma analítica para la predicción de estancia hospitalaria (Length of Stay).")

    if version_modelo == "V4":
        st.caption("Motor Predictivo: Versión 4.0 (Regresor Numérico Exacto en Días)")
    elif version_modelo == "V3":
        st.caption("Motor Predictivo: Versión 3.0 (Alta Resolución con Minería Clínica)")
    elif version_modelo == "V2":
        st.caption("Motor Predictivo: Versión 2.0 (Alta Precisión Clínica)")
    else:
        st.caption("Motor Predictivo: Versión 1.0")

st.markdown("---")

# =========================================================
# 5. NAVEGACIÓN
# =========================================================
selected = option_menu(
    menu_title=None,
    options=["Paciente Individual", "Auditoría Masiva"],
    icons=["person-bounding-box", "bar-chart-fill"],
    menu_icon="cast",
    default_index=0,
    orientation="horizontal",
    styles={
        "container": {
            "padding": "0!important",
            "background-color": "#f8f9fa",
            "border-top": "2px solid #b6b5af",
            "border-bottom": "2px solid #b6b5af"
        },
        "icon": {"color": "#4e6c9f", "font-size": "18px"},
        "nav-link": {
            "font-size": "16px",
            "text-align": "center",
            "margin": "0px",
            "--hover-color": "#e0e0e0",
            "font-family": "Poppins",
            "font-weight": "600",
            "color": "#4e6c9f"
        },
        "nav-link-selected": {
            "background-color": "#253d5b",
            "color": "white"
        },
    }
)

# =========================================================
# PESTAÑA 1: PACIENTE INDIVIDUAL
# =========================================================
if selected == "Paciente Individual":
    st.subheader("Ingreso de Paciente a Urgencias / Piso")

    col1, col2, col3 = st.columns(3)
    with col1:
        edad = st.number_input("Edad del paciente", min_value=0, max_value=120, value=45)
        sexo = st.selectbox("Sexo", ["M", "F"])

    with col2:
        pabellon = st.selectbox(
            "Servicio / Pabellón",
            ["URGENCIAS", "CIRUGIA", "UCI", "HOSPITALIZACION", "TRIAGE"]
        )
        esp = st.text_input("Especialidad (Ej. MEDICINA INTERNA)", "MEDICINA INTERNA").upper().strip()

    with col3:
        dx = st.text_input("Dx Principal CIE-10 (Hasta 4 letras, ej. J450)", "OTR").upper().strip()[:4]

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("**Comorbilidades Adicionales**")

    col_c1, col_c2, col_c3, col_c4, col_c5 = st.columns(5)
    with col_c1:
        diabetes = st.checkbox("Diabetes")
        hipertension = st.checkbox("Hipertensión")
    with col_c2:
        cardiaca = st.checkbox("Enf. Cardíaca")
        epoc = st.checkbox("EPOC")
    with col_c3:
        hemato_onco = st.checkbox("Cáncer / Oncológica")
        quimio = st.checkbox("En Quimioterapia")
    with col_c4:
        hemofilia = st.checkbox("Hemofilia")
        porfiria = st.checkbox("Porfiria")
    with col_c5:
        renal = st.checkbox("Enfermedad Renal")
        vih = st.checkbox("VIH / SIDA")

    st.markdown("<br>", unsafe_allow_html=True)

    if st.button("Calcular Riesgo de Estancia", key="btn_indiv"):
        total_comorbilidades = sum([
            diabetes, hipertension, cardiaca, epoc, hemato_onco,
            quimio, hemofilia, porfiria, renal, vih
        ])

        if version_modelo in ["V2", "V3", "V4"]:
            datos = pd.DataFrame({
                'edad': [edad],
                'Sexo': [sexo],
                'PabellonIngreso': [pabellon],
                'Esp': [esp],
                'Dx_Agrupado': [dx],
                'Diabetes': [int(diabetes)],
                'Hipertension': [int(hipertension)],
                'Cardiaca': [int(cardiaca)],
                'EPOC': [int(epoc)],
                'Hemato_Onco': [int(hemato_onco)],
                'Quimio': [int(quimio)],
                'Hemofilia': [int(hemofilia)],
                'Porfiria': [int(porfiria)],
                'Renal': [int(renal)],
                'VIH': [int(vih)],
                'Total_Comorbilidades': [total_comorbilidades]
            })
        else:
            complejidad_val = 1 if total_comorbilidades > 0 else 0
            datos = pd.DataFrame({
                'edad': [edad],
                'Sexo': [sexo],
                'PabellonIngreso': [pabellon],
                'Esp': [esp],
                'Dx_Agrupado': [dx],
                'Complejidad': [complejidad_val]
            })

        datos_finales = pd.get_dummies(datos).reindex(columns=columnas_entrenamiento, fill_value=0)
        prediccion = modelo.predict(datos_finales)[0]

        st.markdown("---")
        st.subheader("Resultado Clínico")

        if version_modelo == "V4":
            dias_estimados = max(0.5, float(prediccion))

            if dias_estimados > 7:
                st.error(f"🚨 **ALERTA - Riesgo de Ocupación Prolongada**\n\n📍 Estancia estimada: **{dias_estimados:.1f} días**.")
            elif dias_estimados > 3:
                st.warning(f"🟡 **PRECAUCIÓN - Monitoreo estándar requerido**\n\n📍 Estancia estimada: **{dias_estimados:.1f} días**.")
            else:
                st.success(f"🟢 **ÓPTIMO - Flujo rápido esperado**\n\n📍 Estancia estimada: **{dias_estimados:.1f} días**.")
        else:
            pred_texto = str(prediccion)
            if "Corta" in pred_texto:
                st.success(f"{pred_texto} - Flujo rápido esperado. Planear alta temprana.")
            elif "Media" in pred_texto:
                st.warning(f"{pred_texto} - Monitoreo estándar requerido.")
            else:
                st.error(f"{pred_texto} - ALERTA: Alto riesgo de ocupación prolongada. Requiere revisión de caso.")

# =========================================================
# PESTAÑA 2: AUDITORÍA MASIVA
# =========================================================
if selected == "Auditoría Masiva":
    st.subheader("Proyección del Censo Actual")
    st.write("Sube el archivo Excel o CSV del censo matutino para predecir las estancias detectando pacientes desviados.")

    archivo_subido = st.file_uploader("Cargar Censo en formato CSV o Excel", type=["csv", "xlsx", "xls"])

    if archivo_subido is not None:
        try:
            # 1. Lectura
            df_censo = leer_archivo_subido(archivo_subido)
            st.info(f"Se cargaron {len(df_censo)} pacientes del censo.")

            # 2. Estandarización y limpieza
            df_censo = estandarizar_columnas(df_censo)
            df_censo = preparar_pabellones(df_censo)
            df_censo = preparar_edad(df_censo)
            df_censo = preparar_campos_basicos(df_censo)
            df_censo = preparar_dias_actuales(df_censo)

            # 3. Variables del modelo
            df_censo, columnas_x = construir_variables_modelo(df_censo, version_modelo)

            col_disponibles = [c for c in columnas_x if c in df_censo.columns]
            X_masivo = df_censo[col_disponibles].copy()
            X_masivo_encoded = pd.get_dummies(X_masivo).reindex(columns=columnas_entrenamiento, fill_value=0)

            # 4. Predicción
            df_censo['Prediccion_Estancia'] = modelo.predict(X_masivo_encoded)

            # 5. Auditoría
            df_censo, riesgo_prolongado, alertas = auditar_predicciones(df_censo, version_modelo)

            # 6. KPIs
            st.markdown("### 📊 Tablero de Control de Riesgo y Ocupación")

            total_pacientes = len(df_censo)
            desviados = sum(1 for a in alertas if "Desviado" in a)

            kpi1, kpi2, kpi3 = st.columns(3)
            kpi1.metric(label="Pacientes Ingresados", value=total_pacientes)
            kpi2.metric(
                label="Desviados (Superaron Límite Estimado)",
                value=desviados,
                delta=f"{(desviados / total_pacientes) * 100:.1f}%" if total_pacientes > 0 else "0.0%",
                delta_color="inverse"
            )
            kpi3.metric(label="Predicción Estancias Prolongadas", value=riesgo_prolongado)

            # 7. Gráficos
            st.markdown("<br>", unsafe_allow_html=True)
            col_chart1, col_chart2 = st.columns(2)

            with col_chart1:
                if version_modelo == "V4":
                    categorias_orden = [
                        'Estancia Corta (<3 días)',
                        'Estancia Media (3 a 7 días)',
                        'Estancia Prolongada (>7 días)',
                        'Sin predicción'
                    ]
                    df_censo['Categoria_Estancia'] = pd.Categorical(
                        df_censo['Categoria_Estancia'],
                        categories=categorias_orden,
                        ordered=True
                    )

                    riesgo_counts = (
                        df_censo['Categoria_Estancia']
                        .value_counts(dropna=False)
                        .reindex(categorias_orden, fill_value=0)
                        .reset_index()
                    )
                    riesgo_counts.columns = ['Categoría de Estancia', 'Cantidad']
                    riesgo_counts = riesgo_counts[riesgo_counts['Cantidad'] > 0].copy()

                    fig_pie = px.pie(
                        riesgo_counts,
                        values='Cantidad',
                        names='Categoría de Estancia',
                        title='Distribución de Riesgo de Estancia',
                        color='Categoría de Estancia',
                        color_discrete_map={
                            'Estancia Corta (<3 días)': '#4e6c9f',
                            'Estancia Media (3 a 7 días)': '#b6b5af',
                            'Estancia Prolongada (>7 días)': '#253d5b',
                            'Sin predicción': '#d9d9d9'
                        },
                        hole=0.25
                    )
                else:
                    riesgo_counts = df_censo['Prediccion_Estancia'].astype(str).value_counts(dropna=False).reset_index()
                    riesgo_counts.columns = ['Categoría de Estancia', 'Cantidad']

                    fig_pie = px.pie(
                        riesgo_counts,
                        values='Cantidad',
                        names='Categoría de Estancia',
                        title='Distribución de Riesgo de Estancia',
                        hole=0.25
                    )

                fig_pie.update_traces(textposition='inside', textinfo='percent+label')
                fig_pie.update_layout(margin=dict(t=40, b=0, l=0, r=0), legend_title_text='')
                st.plotly_chart(fig_pie, width="stretch")

            with col_chart2:
                desviados_df = df_censo[df_censo['Estado_Auditoria'].str.contains("Desviado", na=False)]
                pabellon_counts = desviados_df['PabellonVisual'].value_counts().reset_index()
                pabellon_counts.columns = ['Pabellón', 'Pacientes Desviados']

                if len(pabellon_counts) > 0:
                    fig_bar = px.bar(
                        pabellon_counts,
                        x='Pabellón',
                        y='Pacientes Desviados',
                        title='Alarmas de Desviación por Pabellón',
                        color_discrete_sequence=['#253d5b']
                    )
                    fig_bar.update_layout(xaxis_tickangle=-45, margin=dict(t=40, b=0, l=0, r=0))
                    st.plotly_chart(fig_bar, width="stretch")
                else:
                    st.info("No hay pacientes desviados para graficar por pabellón.")

            # 8. Radiografía por servicio
            st.markdown("---")
            st.markdown("### 🏥 Radiografía por Servicio Individual (Ubicación Actual / Visual)")

            pabellones = sorted([
                p for p in df_censo['PabellonVisual'].dropna().unique()
                if str(p).strip() != ""
            ])

            if len(pabellones) > 0:
                cols_pab = st.columns(3)

                for i, pab in enumerate(pabellones):
                    df_pab = df_censo[df_censo['PabellonVisual'] == pab]
                    if len(df_pab) == 0:
                        continue

                    estado_counts = df_pab['Estado_Auditoria'].value_counts().reset_index()
                    estado_counts.columns = ['Estado', 'Cant']

                    fig_mini = px.pie(
                        estado_counts,
                        values='Cant',
                        names='Estado',
                        hole=0.5,
                        title=f'{str(pab)[:24]} ({len(df_pab)})',
                        color='Estado',
                        color_discrete_map={
                            'Normal - Dentro del límite estimado': '#b6b5af',
                            'Desviado - Límite Superado': '#253d5b'
                        }
                    )
                    fig_mini.update_traces(textinfo='value')
                    fig_mini.update_layout(
                        showlegend=False,
                        margin=dict(t=30, b=10, l=10, r=10),
                        height=220,
                        title_x=0.5,
                        title_font_size=13
                    )

                    with cols_pab[i % 3]:
                        st.plotly_chart(fig_mini, width="stretch")

            # 9. Detalle nominal
            st.markdown("---")
            st.markdown("### 📋 Detalle Nominal de Auditoría")

            columnas_mostrar = [
                'Identificacion',
                'Paciente_Nombre',
                'Paciente_Apellido',
                'CAMA',
                'edad',
                'PabellonVisual',
                'PabellonIngreso',
                'Esp',
                'Dx_Agrupado',
                'FechaIngreso',
                'Dias_Actuales',
                'Prediccion_Estancia_Texto',
                'Estado_Auditoria'
            ]
            columnas_disponibles = [c for c in columnas_mostrar if c in df_censo.columns]
            df_mostrar = df_censo[columnas_disponibles].copy()

            def resaltar_filas(row):
                if 'Desviado' in str(row.get('Estado_Auditoria', '')):
                    return ['background-color: #ffdce0'] * len(row)
                return [''] * len(row)

            styled_df = df_mostrar.style.apply(resaltar_filas, axis=1).set_table_styles([
                {'selector': 'th', 'props': [('background-color', '#253d5b'), ('color', 'white'), ('font-weight', 'bold')]}
            ])

            st.dataframe(styled_df, width="stretch")

            # 10. Descarga
            excel_data = exportar_excel_estilizado(df_mostrar)

            st.download_button(
                label="📥 Descargar Censo Completo (Excel con Colores)",
                data=excel_data,
                file_name='Censo_Predictivo_Institucional.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                type='primary'
            )

        except Exception as e:
            st.error(f"Error procesando el archivo: {e}")
            st.exception(e)
